# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'anpipy_firstgui.ui'
# Created by: PyQt5 UI code generator 5.15.4
#
# AnPiPy 1.0 - schauerstoff - 24.05.21

import sys
import os
from enum import Enum
from io import BytesIO
import win32clipboard
from PIL import Image
from PyQt5 import QtCore, QtGui, QtWidgets
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import urllib.request

# change paths here
path = "C:\_schauerstoff\japanese\chromedriver.exe"
pic_path = "C:\_schauerstoff\git\AnPiPy\pics"
xlsx_file = '\data\SummerMemories1.xlsx'
worksheet_name = 'summermemories'


class Chrome:
    driver = None
    # last resort: batch to change the name of files while downloading
    batch = 0

    def setup_chrome(self):
        # make error go away :-)
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        # kein chrome fenster wird geoeffnet
        #options.add_argument("--headless")
        self.driver = webdriver.Chrome(path, options=options)
        self.driver.get('https://www.google.de/imghp?hl=de')

    def download_images(self, word):
        self.batch += 1
        self.driver.find_element_by_name('q').clear()
        search_field = self.driver.find_element_by_name('q')
        search_field.send_keys(word)
        search_field.send_keys(Keys.RETURN)
        for i in range(1, 21, 1):
            # str(i) gets the ith pic on page from i = 1 up to i = 20
            img = self.driver.find_element_by_xpath(
                '//*[@id="islrg"]/div[1]/div[' + str(i) + ']/a[1]/div[1]/img')  
            src = img.get_attribute('src')
            name = "pics/Pic" + str(i) + str(self.batch) + ".png"
            urllib.request.urlretrieve(src, name)


class Language(Enum):
    JAP = 1
    ENG = 2
    STK = 3


class Data:
    pos = -1
    lang = Language.JAP
    japanese = None
    english = None

    def setup_data(self):
        # Read xlsx with Kanji in first and English in 4th column
        wb = load_workbook(xlsx_file)
        ws = wb[worksheet_name]
        column = ws['A']
        self.japanese = [column[x].value for x in range(
            1, len(column))]  # 0. entry is "column 1"
        column = ws['D']
        self.english = [column[x].value for x in range(1, len(column))]


class Ui_MainWindow(object):
    all_pics = None
    chrome = Chrome()
    data = Data()
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()

    def setup(self):
        self.data.setup_data()
        self.setup_ui(self.MainWindow)
        self.chrome.setup_chrome()

    def setup_ui(self, MainWindow):
        # region all inits
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1161, 1000)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.gridLayoutWidget.setGeometry(QtCore.QRect(300, 20, 561, 99))
        self.gridLayoutWidget.setObjectName("gridLayoutWidget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.gridLayoutWidget)
        self.gridLayout_2.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.prev = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.prev.setObjectName("prev")
        self.gridLayout_2.addWidget(self.prev, 1, 0, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(
            40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem, 0, 1, 1, 1)
        self.japaneseLab = QtWidgets.QLabel(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setPointSize(30)
        self.japaneseLab.setFont(font)
        self.japaneseLab.setObjectName("japanese")
        self.gridLayout_2.addWidget(self.japaneseLab, 0, 2, 1, 1)
        self.englishLab = QtWidgets.QLabel(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setPointSize(15)
        self.englishLab.setFont(font)
        self.englishLab.setObjectName("english")
        self.gridLayout_2.addWidget(self.englishLab, 1, 2, 1, 1)
        self.next = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.next.setObjectName("next")
        self.gridLayout_2.addWidget(self.next, 1, 4, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(
            40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem1, 0, 3, 1, 1)
        self.gridLayoutWidget_2 = QtWidgets.QWidget(self.centralwidget)
        self.gridLayoutWidget_2.setGeometry(QtCore.QRect(20, 130, 1121, 791))
        self.gridLayoutWidget_2.setObjectName("gridLayoutWidget_2")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.gridLayoutWidget_2)
        self.gridLayout_3.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.Copy16 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy16.setObjectName("Copy16")
        self.gridLayout_3.addWidget(self.Copy16, 7, 3, 1, 1)
        self.pic14 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic14.setObjectName("pic14")
        self.gridLayout_3.addWidget(self.pic14, 6, 1, 1, 1)
        self.Copy13 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy13.setObjectName("Copy13")
        self.gridLayout_3.addWidget(self.Copy13, 7, 0, 1, 1)
        self.Copy7 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy7.setObjectName("Copy7")
        self.gridLayout_3.addWidget(self.Copy7, 3, 2, 1, 1)
        self.Copy5 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy5.setObjectName("Copy5")
        self.gridLayout_3.addWidget(self.Copy5, 3, 0, 1, 1)
        self.Copy9 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy9.setObjectName("Copy9")
        self.gridLayout_3.addWidget(self.Copy9, 5, 0, 1, 1)
        self.pic4 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic4.setObjectName("pic4")
        self.gridLayout_3.addWidget(self.pic4, 0, 3, 1, 1)
        self.Copy6 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy6.setObjectName("Copy6")
        self.gridLayout_3.addWidget(self.Copy6, 3, 1, 1, 1)
        self.pic9 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic9.setObjectName("pic9")
        self.gridLayout_3.addWidget(self.pic9, 4, 0, 1, 1)
        self.Copy15 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy15.setObjectName("Copy15")
        self.gridLayout_3.addWidget(self.Copy15, 7, 2, 1, 1)
        self.pic10 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic10.setObjectName("pic10")
        self.gridLayout_3.addWidget(self.pic10, 4, 1, 1, 1)
        self.pic7 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic7.setObjectName("pic7")
        self.gridLayout_3.addWidget(self.pic7, 2, 2, 1, 1)
        self.pic8 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic8.setObjectName("pic8")
        self.gridLayout_3.addWidget(self.pic8, 2, 3, 1, 1)
        self.pic3 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic3.setObjectName("pic3")
        self.gridLayout_3.addWidget(self.pic3, 0, 2, 1, 1)
        self.pic16 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic16.setObjectName("pic16")
        self.gridLayout_3.addWidget(self.pic16, 6, 3, 1, 1)
        self.pic6 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic6.setObjectName("pic6")
        self.gridLayout_3.addWidget(self.pic6, 2, 1, 1, 1)
        self.Copy12 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy12.setObjectName("Copy12")
        self.gridLayout_3.addWidget(self.Copy12, 5, 3, 1, 1)
        self.Copy11 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy11.setObjectName("Copy11")
        self.gridLayout_3.addWidget(self.Copy11, 5, 2, 1, 1)
        self.pic15 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic15.setObjectName("pic15")
        self.gridLayout_3.addWidget(self.pic15, 6, 2, 1, 1)
        self.Copy10 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy10.setObjectName("Copy10")
        self.gridLayout_3.addWidget(self.Copy10, 5, 1, 1, 1)
        self.pic2 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic2.setObjectName("pic2")
        self.gridLayout_3.addWidget(self.pic2, 0, 1, 1, 1)
        self.pic12 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic12.setObjectName("pic12")
        self.gridLayout_3.addWidget(self.pic12, 4, 3, 1, 1)
        self.pic5 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic5.setObjectName("pic5")
        self.gridLayout_3.addWidget(self.pic5, 2, 0, 1, 1)
        self.pic11 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic11.setObjectName("pic11")
        self.gridLayout_3.addWidget(self.pic11, 4, 2, 1, 1)
        self.Copy1 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy1.setObjectName("Copy1")
        self.gridLayout_3.addWidget(self.Copy1, 1, 0, 1, 1)
        self.Copy14 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy14.setObjectName("Copy14")
        self.gridLayout_3.addWidget(self.Copy14, 7, 1, 1, 1)
        self.Copy4 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy4.setObjectName("Copy4")
        self.gridLayout_3.addWidget(self.Copy4, 1, 3, 1, 1)
        self.Copy3 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy3.setObjectName("Copy3")
        self.gridLayout_3.addWidget(self.Copy3, 1, 2, 1, 1)
        self.pic1 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic1.setObjectName("pic1")
        self.gridLayout_3.addWidget(self.pic1, 0, 0, 1, 1)
        self.Copy2 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy2.setObjectName("Copy2")
        self.gridLayout_3.addWidget(self.Copy2, 1, 1, 1, 1)
        self.Copy8 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.Copy8.setObjectName("Copy8")
        self.gridLayout_3.addWidget(self.Copy8, 3, 3, 1, 1)
        self.pic13 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.pic13.setObjectName("pic13")
        self.gridLayout_3.addWidget(self.pic13, 6, 0, 1, 1)
        self.search_button_S = QtWidgets.QPushButton(self.centralwidget)
        self.search_button_S.setGeometry(QtCore.QRect(890, 70, 131, 28))
        self.search_button_S.setObjectName("search_button_S")
        self.search_button_E = QtWidgets.QPushButton(self.centralwidget)
        self.search_button_E.setGeometry(QtCore.QRect(890, 30, 131, 28))
        self.search_button_E.setObjectName("search_button_E")
        # new
        self.search_button_J = QtWidgets.QPushButton(self.centralwidget)
        self.search_button_J.setGeometry(QtCore.QRect(890, 0, 131, 28))
        self.search_button_J.setObjectName("Japanese")

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1161, 26))
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")
        self.menuAbout = QtWidgets.QMenu(self.menubar)
        self.menuAbout.setObjectName("menuAbout")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionLoad = QtWidgets.QAction(MainWindow)
        self.actionLoad.setObjectName("actionLoad")
        self.actionHelp = QtWidgets.QAction(MainWindow)
        self.actionHelp.setObjectName("actionHelp")
        self.actionStory = QtWidgets.QAction(MainWindow)
        self.actionStory.setObjectName("actionStory")
        self.actionSettings = QtWidgets.QAction(MainWindow)
        self.actionSettings.setObjectName("actionSettings")
        self.menuFile.addAction(self.actionLoad)
        self.menuFile.addAction(self.actionSettings)
        self.menuAbout.addAction(self.actionHelp)
        self.menuAbout.addAction(self.actionStory)
        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuAbout.menuAction())

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # set actions for copy buttons
        self.Copy1.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic1' + str(self.chrome.batch) + '.png'))
        self.Copy2.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic2' + str(self.chrome.batch) + '.png'))
        self.Copy3.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic3' + str(self.chrome.batch) + '.png'))
        self.Copy4.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic4' + str(self.chrome.batch) + '.png'))
        self.Copy5.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic5' + str(self.chrome.batch) + '.png'))
        self.Copy6.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic6' + str(self.chrome.batch) + '.png'))
        self.Copy7.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic7' + str(self.chrome.batch) + '.png'))
        self.Copy8.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic8' + str(self.chrome.batch) + '.png'))
        self.Copy9.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic9' + str(self.chrome.batch) + '.png'))
        self.Copy10.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic10' + str(self.chrome.batch) + '.png'))
        self.Copy11.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic11' + str(self.chrome.batch) + '.png'))
        self.Copy12.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic12' + str(self.chrome.batch) + '.png'))
        self.Copy13.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic13' + str(self.chrome.batch) + '.png'))
        self.Copy14.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic14' + str(self.chrome.batch) + '.png'))
        self.Copy15.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic15' + str(self.chrome.batch) + '.png'))
        self.Copy16.clicked.connect(
            lambda: self.send_to_clipboard('pics/Pic16' + str(self.chrome.batch) + '.png'))
        # self.Copy17.clicked.connect(
        #     lambda: self.send_to_clipboard('pics/Pic17' + str(self.chrome.batch) + '.png'))
        # self.Copy18.clicked.connect(
        #     lambda: self.send_to_clipboard('pics/Pic18' + str(self.chrome.batch) + '.png'))
        # self.Copy19.clicked.connect(
        #     lambda: self.send_to_clipboard('pics/Pic19' + str(self.chrome.batch) + '.png'))
        # self.Copy20.clicked.connect(
        #     lambda: self.send_to_clipboard('pics/Pic20' + str(self.chrome.batch) + '.png'))

        # evtl. noch eine Zeile mehr? self.pic17, self.pic18, self.pic19, self.pic20
        self.all_pics = [self.pic1, self.pic2, self.pic3, self.pic4, self.pic5, self.pic6, self.pic7, self.pic8, self.pic9,
                         self.pic10, self.pic11, self.pic12, self.pic13, self.pic14, self.pic15, self.pic16]

        # Next, Prev
        self.next.clicked.connect(lambda: self.next_word(self.data.pos))
        self.prev.clicked.connect(lambda: self.prev_word(self.data.pos))

        # Language Switch
        self.search_button_E.clicked.connect(
            lambda: self.change_lang(Language.ENG))
        self.search_button_S.clicked.connect(
            lambda: self.change_lang(Language.STK))
        self.search_button_J.clicked.connect(
            lambda: self.change_lang(Language.JAP))

        # endregion

        # exit handler
        self.app.aboutToQuit.connect(self.cleanup)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "AnPiPy 1.0"))
        self.prev.setText(_translate("MainWindow", "<"))
        self.japaneseLab.setText(_translate("MainWindow", "Japanese"))
        self.englishLab.setText(_translate("MainWindow", "English"))
        self.next.setText(_translate("MainWindow", ">"))
        self.Copy16.setText(_translate("MainWindow", "Copy"))
        self.pic14.setText(_translate("MainWindow", "Pic14"))
        self.Copy13.setText(_translate("MainWindow", "Copy"))
        self.Copy7.setText(_translate("MainWindow", "Copy"))
        self.Copy5.setText(_translate("MainWindow", "Copy"))
        self.Copy9.setText(_translate("MainWindow", "Copy"))
        self.pic4.setText(_translate("MainWindow", "Pic4"))
        self.Copy6.setText(_translate("MainWindow", "Copy"))
        self.pic9.setText(_translate("MainWindow", "Pic9"))
        self.Copy15.setText(_translate("MainWindow", "Copy"))
        self.pic10.setText(_translate("MainWindow", "Pic10"))
        self.pic7.setText(_translate("MainWindow", "Pic7"))
        self.pic8.setText(_translate("MainWindow", "Pic8"))
        self.pic3.setText(_translate("MainWindow", "Pic3"))
        self.pic16.setText(_translate("MainWindow", "Pic16"))
        self.pic6.setText(_translate("MainWindow", "Pic6"))
        self.Copy12.setText(_translate("MainWindow", "Copy"))
        self.Copy11.setText(_translate("MainWindow", "Copy"))
        self.pic15.setText(_translate("MainWindow", "Pic15"))
        self.Copy10.setText(_translate("MainWindow", "Copy"))
        self.pic2.setText(_translate("MainWindow", "Pic2"))
        self.pic12.setText(_translate("MainWindow", "Pic12"))
        self.pic5.setText(_translate("MainWindow", "Pic5"))
        self.pic11.setText(_translate("MainWindow", "Pic11"))
        self.Copy1.setText(_translate("MainWindow", "Copy"))
        self.Copy14.setText(_translate("MainWindow", "Copy"))
        self.Copy4.setText(_translate("MainWindow", "Copy"))
        self.Copy3.setText(_translate("MainWindow", "Copy"))
        self.pic1.setText(_translate("MainWindow", "Pic1"))
        self.Copy2.setText(_translate("MainWindow", "Copy"))
        self.Copy8.setText(_translate("MainWindow", "Copy"))
        self.pic13.setText(_translate("MainWindow", "Pic13"))
        self.search_button_S.setText(_translate("MainWindow", "Stock"))
        self.search_button_E.setText(_translate("MainWindow", "English"))
        self.search_button_J.setText(_translate("MainWindow", "Japanese"))
        self.menuFile.setTitle(_translate("MainWindow", "File"))
        self.menuAbout.setTitle(_translate("MainWindow", "About"))
        self.actionLoad.setText(_translate("MainWindow", "Load"))
        self.actionHelp.setText(_translate("MainWindow", "Help"))
        self.actionStory.setText(_translate("MainWindow", "Story"))
        self.actionSettings.setText(_translate("MainWindow", "Settings"))

    def load_word(self, pos):
        j_word = self.data.japanese[pos]
        e_word = self.data.english[pos]
        self.japaneseLab.setText(j_word)
        self.englishLab.setText(e_word)
        if(self.data.lang == Language.JAP):
            self.chrome.download_images(j_word)
            self.fill_images()
        elif(self.data.lang == Language.ENG):
            self.chrome.download_images(e_word)
            self.fill_images()
        else:
            e_word += " stock photo"
            self.chrome.download_images(e_word)
            self.fill_images()

    def next_word(self, pos):
        # self.data.lang = Language.JAP
        if (pos < len(self.data.japanese)):  # -1?
            self.data.pos += 1
        self.load_word(self.data.pos)

    def prev_word(self, pos):
        # self.data.lang = Language.JAP
        if (pos > 0):
            self.data.pos -= 1
            self.load_word(self.data.pos)

    def change_lang(self, lang):
        if not isinstance(lang, Language):
            raise TypeError('direction must be an instance of Language Enum')
        self.data.lang = lang
        self.load_word(self.data.pos)

    def fill_images(self):
        for i in range(len(self.all_pics)):
            path = "pics/Pic" + str(i+1) + str(self.chrome.batch)
            self.all_pics[i].setPixmap(QtGui.QPixmap(path))
        # self.app.processEvents()  # process all the events present on the queue, not needed afterall

    def send_to_clipboard(self, image_path):
        image = Image.open(image_path)
        output = BytesIO()
        image.convert('RGB').save(output, 'BMP')
        data = output.getvalue()[14:]
        output.close()
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
        win32clipboard.CloseClipboard()

    def cleanup(self):
        # delete all images in folder /pics
        for file in os.scandir(pic_path):
            if file.name.endswith(".png"):
                os.unlink(file.path)


if __name__ == "__main__":
    ui = Ui_MainWindow()
    ui.setup()
    ui.MainWindow.show()
    sys.exit(ui.app.exec_())
