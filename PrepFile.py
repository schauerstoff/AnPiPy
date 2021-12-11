import openpyxl as pyxl
from bs4 import BeautifulSoup
from pprint import pprint
from PyInquirer import prompt, Separator
from examples import custom_style_1

#How to use:
# Export Notes as plain text > txt from Anki
# Open in Excel, change paths
# Run and choose words
# Open in Excel, repair, export as csv
# Copy content in nre UTF-8 encoded csv
# Import in Anki

# change paths here
xlsx_file = 'SummerMemories1.xlsx'
worksheet_name = 'summermemories'

# set columns
kanji_index = 1
word_index = 4
resource_index = 6

# set new resource name
resource_name = "Return to Shironagasu Island"


class Data:
    # setup
    def __init__(self):
        self.wb = pyxl.load_workbook(xlsx_file)
        self.ws = self.wb[worksheet_name]
        self.html = []

    def set_resource(self):
        #resource = self.ws[resource_index]
        #print(len(resource))
        for x in range(2, self.ws.max_row + 1):  # 0. entry is "column x"
            self.ws.cell(row=x, column=resource_index).value = resource_name
            # print(self.ws.cell(row=x, column=6).value)
        self.wb.save(xlsx_file)

    def remove_tags(self):
        #words = self.ws[word_index]
        for x in range(2, self.ws.max_row + 1):  # 0. entry is "column x"
            s = str(self.ws.cell(row=x, column=word_index).value)
            self.ws.cell(row=x, column=word_index).value = s.replace(
                '<div style="text-align: left;">', '').replace('</div>', '').replace('<br>', '').replace('</br>', '')
        self.wb.save(xlsx_file)

    def write(self, string, row_index):
        self.ws.cell(row=row_index, column=word_index).value = string
        self.wb.save(xlsx_file)

    def flatten_and_seperate(self, wordlist):
        # if 2D, flatten 2D-list and add seperator between elements
        if(isinstance(wordlist[0], list)):
            return ', '.join([j for sub in wordlist for j in sub])
        else:
            return wordlist[0]

    def get_html_soup(self):
        #words = self.ws[word_index]
        #print(len(words))
        for x in range(2, self.ws.max_row + 1):  # 0. entry is "column x"
            self.html.append(BeautifulSoup(self.ws.cell(
                row=x, column=word_index).value, features="lxml"))
        return self.html

    # Returns each li from ol{ul{li}}/ol{li}/p{}, PyInquirer.separator between each ol-element, as a list
    # CANNOT handle the use of ' in strings. will remove them.
    def parse_html_list(self, html_code):
        res = []
        olList = html_code.find_all('ul')
        if(olList):
            for elem in olList:
                soup = BeautifulSoup(str(elem), features="lxml")
                ulList = soup.find_all('li')
                for word in ulList:
                    word = str(word).replace('<li>', '').replace(
                        '</li>', '').replace('\'', '').strip()
                    res.append(word)
                res.append(Separator('___________________'))
            return res
        # no nested ul-List, check only one ol
        else:
            oneList = html_code.find_all('li')
            if(oneList):
                for word in oneList:
                    word = str(word).replace('<li>', '').replace(
                        '</li>', '').replace('\'', '').strip()
                    res.append(word)
                return res
            # only one word!
            else:
                oneWord = html_code.find_all('p')
                for word in oneWord:

                    word = str(word).replace('<p>', '').replace(
                        '</p>', '').replace('\'', '').strip()
                    res.append(word)
                return res


class CLI:
    # setup
    def __init__(self):
        self.choices = []

    # fills according to pyinquirer question array format
    def fill_choices_array(self, word_list):
        # remove duplicates
        word_list = list(dict.fromkeys(word_list))
        self.choices = []
        for word in word_list:
            if(not type(word) == Separator):
                self.choices.append({'name': word})
            else:
                self.choices.append(word)

    # validation doesnt work for checkboxes :(
    def user_select(self, kanji):
        # if not empty
        if(len(self.choices) > 1):
            questions = [
                {
                    'type': 'checkbox',
                    'qmark': '',
                    'message': 'Select translations for ' + kanji,
                    'name': 'topp',
                    'choices': self.choices,
                    'validate': lambda answer: 'You must choose at least one word.'
                    if len(answer) == 0 else True
                }
            ]
            answers = prompt(questions, style=custom_style_1)
            pprint(list(answers.values()))
            return list(answers.values())
        else:
            print([self.choices[0].get('name')])
            return [self.choices[0].get('name')]


if __name__ == "__main__":
    # 1. Clean data
    data = Data()
    data.set_resource()
    data.remove_tags()

    # 2. Oeffne xlsx Datei und repariere sie. Bestaetige. (Not needed!!)
    #username = input("Open xlsx, repair, close xlsx and confirm with Enter")
    # 3. Edit data by choosing words
    cli = CLI()

    i = 2
    html_soup = data.get_html_soup()
    for html in html_soup:
        html_choices = data.parse_html_list(html)
        cli.fill_choices_array(html_choices)
        kanji = data.ws.cell(row=i, column=kanji_index).value
        answer = cli.user_select(kanji)
        data.write(data.flatten_and_seperate(answer), i)
        i += 1
