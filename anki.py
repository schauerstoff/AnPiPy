from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time
import urllib.request
PATH = "C:\Program Files (x86)\chromedriver.exe"


# make error go away :-)
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])

driver = webdriver.Chrome(PATH, options=options)
# driver.get('https://www.google.de/imghp?hl=de')
driver.get('https://www.google.com/search?q=selenium+copy+image+selected+from+google+images&client=firefox-b-d&source=lnms&tbm=isch&sa=X&ved=2ahUKEwjw7dTOltPwAhWBzqQKHdMJDLkQ_AUoAnoECAEQBA&biw=1920&bih=978&dpr=1.25#imgrc=QJCXoBknYJevgM')
# try to go by id or name bc they will be unique (most likely)
# search_field = driver.find_element_by_name('q') #we dont do that here

# Read in xlsx with Kanji in first and English in 4th row
wb = load_workbook("Mappe1.xlsx")  # Work Book
ws = wb['sheet1']  # Work Sheet wb.get_sheet_by_name('sheet1')
column = ws['A']  # Column
japanese = [column[x].value for x in range(
    1, len(column))]  # 0. entry is "column 1"
column = ws['D']  # Column
english = [column[x].value for x in range(1, len(column))]

# for i in range(len(japanese)):
#     # muss vor jeder interaktion neu definiert werden
#     driver.find_element_by_name('q').clear()
#     search_field = driver.find_element_by_name('q')
#     search_field.send_keys(japanese[i])
#     search_field.send_keys(Keys.RETURN)

#     response = input('English: e >')
#     if response == 'e':
#         driver.find_element_by_name('q').clear()
#         search_field = driver.find_element_by_name('q')
#         search_field.send_keys(english[i])
#         search_field.send_keys(Keys.RETURN)
#         driver.find_element_by_name('q').clear()
#         response = input('anything >')
# time.sleep(2)
# print(japanese[i])
# print(english[i])

# driver.quit() #oder close()?

# retrieve the picture
# img class="n3VNCb" für das aktive Bild

driver.find_element_by_name('q').clear()
search_field = driver.find_element_by_name('q')
# search_field.send_keys(japanese[i])
# search_field.send_keys(Keys.RETURN)

# get the image source
img = driver.find_element_by_xpath(
    '//*[@id="islrg"]/div[1]/div[' + str(1) + ']/a[1]/div[1]/img')  # str(i) gets the ith pic on page from i = 1 up to i = 20
src = img.get_attribute('src')
time.sleep(2)
# download the image
urllib.request.urlretrieve(src, "captcha.png")
