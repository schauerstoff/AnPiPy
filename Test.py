import openpyxl as pyxl

wb = pyxl.load_workbook("SummerMemories1.xlsx")
ws = wb["summermemories"]

resource = ws['A']
print(len(resource))
for x in range(2, ws.max_row):  
    print(ws.cell(row=x, column=1).value)